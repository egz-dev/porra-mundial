import openpyxl
wb = openpyxl.load_workbook('mock/Porra Mundial 2026.xlsx')
print('Sheets:', wb.sheetnames)

# Check participants
ws = wb['Respuestas']
headers = [str(v) for v in next(ws.iter_rows(values_only=True))]
print(f'\nHeaders: {headers}')

part_count = 0
provinces = {}
for row in ws.iter_rows(values_only=True):
    if row[0] is None:
        continue
    part_count += 1
    prov = str(row[3]) if row[3] else '?'
    provinces[prov] = provinces.get(prov, 0) + 1

print(f'\nRespuestas: {part_count} participants')
print(f'Provinces ({len(provinces)}):')
for p, count in sorted(provinces.items()):
    print(f'  {p}: {count}')

# Check resultados
ws2 = wb['Resultados']
match_count = 0
ft_count = 0
ns_count = 0
group_count = 0
ko_counts = {}
for row in ws2.iter_rows(values_only=True):
    if row[0] is None or str(row[0]) == 'matchId':
        continue
    match_count += 1
    status = str(row[5]) if row[5] else '?'
    fase = str(row[6]) if row[6] else '?'
    if status == 'FT':
        ft_count += 1
    elif status == 'NS':
        ns_count += 1
    ko_counts[fase] = ko_counts.get(fase, 0) + 1

print(f'\nResultados: {match_count} matches')
print(f'  FT: {ft_count}')
print(f'  NS: {ns_count}')
print(f'  By phase:')
for phase in ['group','r32','r16','qf','sf','3rd','final']:
    print(f'    {phase}: {ko_counts.get(phase, 0)}')

# Verify first knockout rows have teams
print('\nFirst 5 knockout matches:')
for row in ws2.iter_rows(values_only=True):
    fase = str(row[6]) if row[6] else ''
    if fase == 'r32' and row[1] and row[2]:
        print(f'  {row[0]}: {row[1]} vs {row[2]} -> {row[3]}-{row[4]} ({row[5]})')

# Check 3rd place and final
print('\nFinal & 3rd place:')
for row in ws2.iter_rows(values_only=True):
    fase = str(row[6]) if row[6] else ''
    if fase in ['3rd', 'final']:
        print(f'  {fase}: {row[1]} vs {row[2]} -> {row[3]}-{row[4]} ({row[5]})')

wb.close()
print('\n✅ Verification complete')
