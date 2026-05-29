import openpyxl
wb = openpyxl.load_workbook('mock/Porra Mundial 2026.xlsx')

print('=== RESPUESTAS ===')
ws = wb['Respuestas']
for i, row in enumerate(ws.iter_rows(values_only=True)):
    vals = [str(v) if v is not None else 'NULL' for v in row]
    print(f'R{i}: {" ||| ".join(vals)}')

print('\n=== RESULTADOS ===')
ws2 = wb['Resultados']
for i, row in enumerate(ws2.iter_rows(values_only=True)):
    vals = [str(v) if v is not None else 'NULL' for v in row]
    print(f'R{i}: {" | ".join(vals)}')
