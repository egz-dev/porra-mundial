"""
Generate a comprehensive mock Excel for Porra Mundial 2026.
- 26 participants across 9 provinces
- 72 group stage matches with FT results
- 32 knockout matches with results consistent with group standings
"""
import openpyxl
from datetime import datetime, timedelta

EXCEL_PATH = 'mock/Porra Mundial 2026.xlsx'

# ── Teams by group (from gruposMundial.js) ─────────────────
GRUPOS_MUNDIAL = {
    'A': ['México', 'Sudáfrica', 'Corea del Sur', 'República Checa'],
    'B': ['Canadá', 'Bosnia-Herzegovina', 'Catar', 'Suiza'],
    'C': ['Brasil', 'Marruecos', 'Haití', 'Escocia'],
    'D': ['EEUU', 'Paraguay', 'Australia', 'Turquía'],
    'E': ['Alemania', 'Curaçao', 'Costa de Marfil', 'Ecuador'],
    'F': ['Países Bajos', 'Japón', 'Suecia', 'Túnez'],
    'G': ['Bélgica', 'Egipto', 'Irán', 'Nueva Zelanda'],
    'H': ['España', 'Cabo Verde', 'Arabia Saudí', 'Uruguay'],
    'I': ['Francia', 'Senegal', 'Irak', 'Noruega'],
    'J': ['Argentina', 'Argelia', 'Austria', 'Jordania'],
    'K': ['Portugal', 'Rep. Dem. del Congo', 'Uzbekistán', 'Colombia'],
    'L': ['Inglaterra', 'Croacia', 'Ghana', 'Panamá'],
}

# ── Deterministic group match results ─────────────────────
# 72 matches: 6 per group × 12 groups
# Match order within each group: (1v2, 3v4), (1v3, 2v4), (1v4, 2v3)
# Results biased toward stronger teams based on fictional FIFA ranking

GROUP_MATCHES = {
    # Grupo A
    'A01': ('México', 'Sudáfrica', 3, 0, 0, 0, '2026-06-11T16:00:00Z'),
    'A02': ('Corea del Sur', 'República Checa', 1, 0, 0, 0, '2026-06-11T19:00:00Z'),
    'A03': ('México', 'Corea del Sur', 2, 0, 0, 0, '2026-06-16T17:00:00Z'),
    'A04': ('Sudáfrica', 'República Checa', 0, 2, 0, 0, '2026-06-16T20:00:00Z'),
    'A05': ('México', 'República Checa', 2, 0, 0, 1, '2026-06-21T18:00:00Z'),  # Checa roja
    'A06': ('Sudáfrica', 'Corea del Sur', 0, 1, 0, 0, '2026-06-21T18:00:00Z'),
    # Grupo B
    'B01': ('Canadá', 'Bosnia-Herzegovina', 2, 0, 0, 0, '2026-06-12T15:00:00Z'),
    'B02': ('Catar', 'Suiza', 0, 3, 0, 0, '2026-06-12T18:00:00Z'),
    'B03': ('Canadá', 'Catar', 1, 0, 0, 0, '2026-06-17T16:00:00Z'),
    'B04': ('Bosnia-Herzegovina', 'Suiza', 0, 1, 0, 0, '2026-06-17T19:00:00Z'),
    'B05': ('Canadá', 'Suiza', 1, 1, 0, 0, '2026-06-22T17:00:00Z'),
    'B06': ('Bosnia-Herzegovina', 'Catar', 0, 0, 0, 0, '2026-06-22T17:00:00Z'),
    # Grupo C
    'C01': ('Brasil', 'Marruecos', 4, 0, 0, 0, '2026-06-13T17:00:00Z'),
    'C02': ('Haití', 'Escocia', 1, 2, 0, 0, '2026-06-13T20:00:00Z'),
    'C03': ('Brasil', 'Haití', 5, 0, 0, 0, '2026-06-18T18:00:00Z'),
    'C04': ('Marruecos', 'Escocia', 1, 0, 0, 1, '2026-06-18T21:00:00Z'),  # Marruecos roja
    'C05': ('Brasil', 'Escocia', 4, 0, 0, 0, '2026-06-23T16:00:00Z'),
    'C06': ('Marruecos', 'Haití', 2, 0, 0, 0, '2026-06-23T16:00:00Z'),
    # Grupo D
    'D01': ('EEUU', 'Paraguay', 1, 1, 0, 0, '2026-06-13T15:00:00Z'),
    'D02': ('Australia', 'Turquía', 0, 2, 0, 0, '2026-06-13T18:00:00Z'),
    'D03': ('EEUU', 'Australia', 2, 0, 0, 0, '2026-06-18T17:00:00Z'),
    'D04': ('Paraguay', 'Turquía', 0, 1, 0, 1, '2026-06-18T20:00:00Z'),  # Paraguay roja
    'D05': ('EEUU', 'Turquía', 0, 1, 0, 0, '2026-06-23T18:00:00Z'),
    'D06': ('Paraguay', 'Australia', 2, 0, 0, 0, '2026-06-23T18:00:00Z'),
    # Grupo E
    'E01': ('Alemania', 'Curaçao', 6, 0, 0, 0, '2026-06-14T16:00:00Z'),
    'E02': ('Costa de Marfil', 'Ecuador', 1, 2, 0, 0, '2026-06-14T19:00:00Z'),
    'E03': ('Alemania', 'Costa de Marfil', 4, 0, 0, 0, '2026-06-19T17:00:00Z'),
    'E04': ('Curaçao', 'Ecuador', 0, 3, 0, 1, '2026-06-19T20:00:00Z'),  # Curaçao roja
    'E05': ('Alemania', 'Ecuador', 3, 1, 0, 0, '2026-06-24T18:00:00Z'),
    'E06': ('Curaçao', 'Costa de Marfil', 1, 2, 0, 0, '2026-06-24T18:00:00Z'),
    # Grupo F
    'F01': ('Países Bajos', 'Japón', 3, 0, 0, 0, '2026-06-14T15:00:00Z'),
    'F02': ('Suecia', 'Túnez', 1, 0, 0, 0, '2026-06-14T18:00:00Z'),
    'F03': ('Países Bajos', 'Suecia', 4, 0, 0, 0, '2026-06-19T16:00:00Z'),
    'F04': ('Japón', 'Túnez', 2, 0, 0, 0, '2026-06-19T19:00:00Z'),
    'F05': ('Países Bajos', 'Túnez', 5, 0, 0, 0, '2026-06-24T16:00:00Z'),
    'F06': ('Japón', 'Suecia', 2, 0, 0, 0, '2026-06-24T16:00:00Z'),
    # Grupo G
    'G01': ('Bélgica', 'Egipto', 2, 0, 0, 0, '2026-06-15T16:00:00Z'),
    'G02': ('Irán', 'Nueva Zelanda', 0, 1, 0, 0, '2026-06-15T19:00:00Z'),
    'G03': ('Bélgica', 'Irán', 3, 0, 0, 0, '2026-06-20T17:00:00Z'),
    'G04': ('Egipto', 'Nueva Zelanda', 2, 0, 0, 0, '2026-06-20T20:00:00Z'),
    'G05': ('Bélgica', 'Nueva Zelanda', 4, 0, 0, 0, '2026-06-25T18:00:00Z'),
    'G06': ('Egipto', 'Irán', 2, 0, 0, 0, '2026-06-25T18:00:00Z'),
    # Grupo H
    'H01': ('España', 'Cabo Verde', 4, 0, 0, 0, '2026-06-15T15:00:00Z'),
    'H02': ('Arabia Saudí', 'Uruguay', 0, 2, 0, 0, '2026-06-15T18:00:00Z'),
    'H03': ('España', 'Arabia Saudí', 5, 0, 0, 0, '2026-06-20T16:00:00Z'),
    'H04': ('Cabo Verde', 'Uruguay', 0, 3, 0, 0, '2026-06-20T19:00:00Z'),
    'H05': ('España', 'Uruguay', 3, 0, 0, 0, '2026-06-25T16:00:00Z'),
    'H06': ('Cabo Verde', 'Arabia Saudí', 1, 1, 0, 0, '2026-06-25T16:00:00Z'),
    # Grupo I
    'I01': ('Francia', 'Senegal', 2, 0, 0, 0, '2026-06-16T15:00:00Z'),
    'I02': ('Irak', 'Noruega', 0, 1, 0, 0, '2026-06-16T18:00:00Z'),
    'I03': ('Francia', 'Irak', 3, 0, 0, 0, '2026-06-21T16:00:00Z'),
    'I04': ('Senegal', 'Noruega', 1, 0, 0, 1, '2026-06-21T19:00:00Z'),  # Noruega roja
    'I05': ('Francia', 'Noruega', 2, 1, 0, 0, '2026-06-26T17:00:00Z'),
    'I06': ('Senegal', 'Irak', 2, 0, 0, 0, '2026-06-26T17:00:00Z'),
    # Grupo J
    'J01': ('Argentina', 'Argelia', 2, 0, 0, 0, '2026-06-17T15:00:00Z'),
    'J02': ('Austria', 'Jordania', 2, 0, 0, 0, '2026-06-17T18:00:00Z'),
    'J03': ('Argentina', 'Austria', 1, 0, 0, 0, '2026-06-22T16:00:00Z'),
    'J04': ('Argelia', 'Jordania', 1, 0, 0, 1, '2026-06-22T19:00:00Z'),  # Jordania roja
    'J05': ('Argentina', 'Jordania', 4, 0, 0, 0, '2026-06-26T19:00:00Z'),
    'J06': ('Argelia', 'Austria', 0, 1, 0, 0, '2026-06-26T19:00:00Z'),
    # Grupo K
    'K01': ('Portugal', 'Rep. Dem. del Congo', 1, 0, 0, 1, '2026-06-12T17:00:00Z'),  # Congo roja
    'K02': ('Uzbekistán', 'Colombia', 0, 3, 0, 0, '2026-06-12T20:00:00Z'),
    'K03': ('Portugal', 'Uzbekistán', 3, 0, 0, 0, '2026-06-17T17:00:00Z'),
    'K04': ('Rep. Dem. del Congo', 'Colombia', 1, 2, 0, 0, '2026-06-17T20:00:00Z'),
    'K05': ('Portugal', 'Colombia', 0, 1, 0, 0, '2026-06-22T18:00:00Z'),
    'K06': ('Rep. Dem. del Congo', 'Uzbekistán', 2, 0, 0, 0, '2026-06-22T18:00:00Z'),
    # Grupo L
    'L01': ('Inglaterra', 'Croacia', 2, 0, 0, 0, '2026-06-14T15:00:00Z'),
    'L02': ('Ghana', 'Panamá', 2, 0, 0, 0, '2026-06-14T18:00:00Z'),
    'L03': ('Inglaterra', 'Ghana', 3, 0, 0, 0, '2026-06-18T16:00:00Z'),
    'L04': ('Croacia', 'Panamá', 3, 0, 0, 0, '2026-06-18T19:00:00Z'),
    'L05': ('Inglaterra', 'Panamá', 5, 0, 0, 0, '2026-06-23T17:00:00Z'),
    'L06': ('Croacia', 'Ghana', 1, 0, 0, 0, '2026-06-23T17:00:00Z'),
}

# ── Group match IDs ────────────────────────────────────────
GROUP_MATCH_IDS = {
    'A': [537327, 537328, 537330, 537331, 537332, 537333],
    'B': [537334, 537335, 537336, 537337, 537338, 537339],
    'C': [537340, 537341, 537342, 537343, 537344, 537345],
    'D': [537346, 537347, 537348, 537349, 537350, 537351],
    'E': [537352, 537353, 537354, 537355, 537356, 537357],
    'F': [537358, 537359, 537360, 537361, 537362, 537363],
    'G': [537364, 537365, 537366, 537367, 537368, 537369],
    'H': [537370, 537371, 537372, 537373, 537374, 537375],
    'I': [537376, 537377, 537378, 537379, 537380, 537381],
    'J': [537382, 537383, 537384, 537385, 537386, 537387],
    'K': [537388, 537389, 537390, 537391, 537392, 537393],
    'L': [537394, 537395, 537396, 537397, 537398, 537399],
}

# ── Original 11 participants (reconstructed) ───────────────
ORIGINAL_PARTICIPANTS = [
    ('Edu García', 'Ciudad Real', 'edu.garcia@email.com',
     ['3 - Argentina 🇦🇷', '6 - Países Bajos 🇳🇱', '7 - Alemania 🇩🇪'],
     ['12 - Ecuador 🇪🇨', '13 - Canadá 🇨🇦', '14 - Suiza 🇨🇭', '15 - Austria 🇦🇹'],
     ['26 - Corea del Sur 🇰🇷', '27 - Egipto 🇪🇬', '28 - Australia 🇦🇺'],
     ['38 - Nueva Zelanda 🇳🇿'],
     ['39 - Curaçao 🇨🇼', '40 - Jordania 🇯🇴']),
    ('Ana López', 'Albacete', 'ana.lopez@email.com',
     ['2 - España 🇪🇸', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '8 - Portugal 🇵🇹'],
     [],
     ['17 - Senegal 🇸🇳', '18 - Japón 🇯🇵', '20 - EEUU 🇺🇸', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '34 - Ghana 🇬🇭', '35 - Irán 🇮🇷'],
     ['41 - Sudáfrica 🇿🇦', '45 - Irak 🇮🇶']),
    ('Carlos Ruiz', 'Toledo', 'carlos.ruiz@email.com',
     ['1 - Brasil 🇧🇷', '5 - Francia 🇫🇷', '7 - Alemania 🇩🇪'],
     ['9 - Noruega 🇳🇴', '10 - Bélgica 🇧🇪', '11 - Colombia 🇨🇴'],
     ['17 - Senegal 🇸🇳', '25 - Marruecos 🇲🇦', '28 - Australia 🇦🇺'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '36 - Uzbekistán 🇺🇿', '38 - Nueva Zelanda 🇳🇿'],
     ['44 - Catar 🇶🇦', '46 - Panamá 🇵🇦']),
    ('Marta Pérez', 'Ciudad Real', 'marta.perez@email.com',
     ['2 - España 🇪🇸', '6 - Países Bajos 🇳🇱', '8 - Portugal 🇵🇹'],
     ['11 - Colombia 🇨🇴', '14 - Suiza 🇨🇭', '15 - Austria 🇦🇹', '16 - México 🇲🇽'],
     ['20 - EEUU 🇺🇸', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾'],
     ['37 - Costa de Marfil 🇨🇮', '38 - Nueva Zelanda 🇳🇿'],
     ['39 - Curaçao 🇨🇼', '42 - Arabia Saudí 🇸🇦']),
    ('Javier Sánchez', 'Albacete', 'javier.sanchez@email.com',
     ['3 - Argentina 🇦🇷', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '5 - Francia 🇫🇷'],
     ['9 - Noruega 🇳🇴', '10 - Bélgica 🇧🇪', '11 - Colombia 🇨🇴'],
     ['17 - Senegal 🇸🇳', '18 - Japón 🇯🇵', '25 - Marruecos 🇲🇦'],
     ['34 - Ghana 🇬🇭', '37 - Costa de Marfil 🇨🇮', '38 - Nueva Zelanda 🇳🇿'],
     ['41 - Sudáfrica 🇿🇦', '44 - Catar 🇶🇦']),
    ('Lucía Martínez', 'Toledo', 'lucia.martinez@email.com',
     ['1 - Brasil 🇧🇷', '6 - Países Bajos 🇳🇱', '7 - Alemania 🇩🇪'],
     ['10 - Bélgica 🇧🇪', '12 - Ecuador 🇪🇨', '14 - Suiza 🇨🇭', '16 - México 🇲🇽'],
     ['17 - Senegal 🇸🇳', '20 - EEUU 🇺🇸', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '38 - Nueva Zelanda 🇳🇿'],
     ['39 - Curaçao 🇨🇼']),
    ('Laura Díaz', 'Cuenca', 'laura.diaz@email.com',
     ['3 - Argentina 🇦🇷', '5 - Francia 🇫🇷', '7 - Alemania 🇩🇪'],
     ['9 - Noruega 🇳🇴', '11 - Colombia 🇨🇴', '14 - Suiza 🇨🇭', '15 - Austria 🇦🇹'],
     ['20 - EEUU 🇺🇸', '25 - Marruecos 🇲🇦', '28 - Australia 🇦🇺'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '34 - Ghana 🇬🇭'],
     ['40 - Jordania 🇯🇴', '42 - Arabia Saudí 🇸🇦']),
    ('Pablo Fernández', 'Ciudad Real', 'pablo.fernandez@email.com',
     ['2 - España 🇪🇸', '8 - Portugal 🇵🇹'],
     ['10 - Bélgica 🇧🇪', '11 - Colombia 🇨🇴', '13 - Canadá 🇨🇦', '15 - Austria 🇦🇹'],
     ['17 - Senegal 🇸🇳', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾', '28 - Australia 🇦🇺'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '37 - Costa de Marfil 🇨🇮'],
     ['46 - Panamá 🇵🇦']),
    ('Sara Moreno', 'Albacete', 'sara.moreno@email.com',
     ['1 - Brasil 🇧🇷', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '6 - Países Bajos 🇳🇱'],
     ['9 - Noruega 🇳🇴', '12 - Ecuador 🇪🇨', '16 - México 🇲🇽'],
     ['18 - Japón 🇯🇵', '20 - EEUU 🇺🇸', '23 - Croacia 🇭🇷', '25 - Marruecos 🇲🇦'],
     ['34 - Ghana 🇬🇭', '35 - Irán 🇮🇷', '38 - Nueva Zelanda 🇳🇿'],
     ['39 - Curaçao 🇨🇼']),
    ('David Serrano', 'Cuenca', 'david.serrano@email.com',
     ['2 - España 🇪🇸', '3 - Argentina 🇦🇷', '7 - Alemania 🇩🇪'],
     ['10 - Bélgica 🇧🇪', '14 - Suiza 🇨🇭', '15 - Austria 🇦🇹'],
     ['17 - Senegal 🇸🇳', '20 - EEUU 🇺🇸', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '37 - Costa de Marfil 🇨🇮'],
     ['40 - Jordania 🇯🇴', '41 - Sudáfrica 🇿🇦']),
    ('Elena Torres', 'Toledo', 'elena.torres@email.com',
     ['1 - Brasil 🇧🇷', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '5 - Francia 🇫🇷'],
     ['9 - Noruega 🇳🇴', '11 - Colombia 🇨🇴', '16 - México 🇲🇽'],
     ['18 - Japón 🇯🇵', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾', '28 - Australia 🇦🇺'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '34 - Ghana 🇬🇭', '35 - Irán 🇮🇷'],
     ['39 - Curaçao 🇨🇼']),
]

# ── New participants (15, = 13 teams each) ─────────────────
NEW_PARTICIPANTS = [
    ('Alejandro Moreno', 'Madrid', 'alejandro.moreno@email.com',
     ['1 - Brasil 🇧🇷', '2 - España 🇪🇸', '5 - Francia 🇫🇷'],
     ['10 - Bélgica 🇧🇪', '11 - Colombia 🇨🇴'],
     ['17 - Senegal 🇸🇳', '18 - Japón 🇯🇵', '25 - Marruecos 🇲🇦', '28 - Australia 🇦🇺'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '34 - Ghana 🇬🇭'],
     ['39 - Curaçao 🇨🇼', '44 - Catar 🇶🇦']),
    ('Carmen Ruiz', 'Madrid', 'carmen.ruiz@email.com',
     ['3 - Argentina 🇦🇷', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '6 - Países Bajos 🇳🇱'],
     ['9 - Noruega 🇳🇴', '14 - Suiza 🇨🇭'],
     ['20 - EEUU 🇺🇸', '21 - Turquía 🇹🇷', '23 - Croacia 🇭🇷', '24 - Uruguay 🇺🇾'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '36 - Uzbekistán 🇺🇿'],
     ['42 - Arabia Saudí 🇸🇦', '45 - Irak 🇮🇶']),
    ('Marc Torres', 'Barcelona', 'marc.torres@email.com',
     ['1 - Brasil 🇧🇷', '3 - Argentina 🇦🇷', '7 - Alemania 🇩🇪'],
     ['10 - Bélgica 🇧🇪', '11 - Colombia 🇨🇴'],
     ['17 - Senegal 🇸🇳', '20 - EEUU 🇺🇸', '25 - Marruecos 🇲🇦', '28 - Australia 🇦🇺'],
     ['34 - Ghana 🇬🇭', '38 - Nueva Zelanda 🇳🇿'],
     ['39 - Curaçao 🇨🇼', '46 - Panamá 🇵🇦']),
    ('Laia Serra', 'Barcelona', 'laia.serra@email.com',
     ['2 - España 🇪🇸', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '5 - Francia 🇫🇷'],
     ['9 - Noruega 🇳🇴', '10 - Bélgica 🇧🇪'],
     ['18 - Japón 🇯🇵', '21 - Turquía 🇹🇷', '23 - Croacia 🇭🇷', '24 - Uruguay 🇺🇾'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '37 - Costa de Marfil 🇨🇮'],
     ['41 - Sudáfrica 🇿🇦', '44 - Catar 🇶🇦']),
    ('Vicent Navarro', 'Valencia', 'vicent.navarro@email.com',
     ['3 - Argentina 🇦🇷', '6 - Países Bajos 🇳🇱', '8 - Portugal 🇵🇹'],
     ['11 - Colombia 🇨🇴', '12 - Ecuador 🇪🇨'],
     ['17 - Senegal 🇸🇳', '20 - EEUU 🇺🇸', '23 - Croacia 🇭🇷', '25 - Marruecos 🇲🇦'],
     ['34 - Ghana 🇬🇭', '38 - Nueva Zelanda 🇳🇿'],
     ['40 - Jordania 🇯🇴', '42 - Arabia Saudí 🇸🇦']),
    ('María Ferrer', 'Valencia', 'maria.ferrer@email.com',
     ['1 - Brasil 🇧🇷', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '7 - Alemania 🇩🇪'],
     ['9 - Noruega 🇳🇴', '13 - Canadá 🇨🇦'],
     ['18 - Japón 🇯🇵', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾', '28 - Australia 🇦🇺'],
     ['35 - Irán 🇮🇷', '37 - Costa de Marfil 🇨🇮'],
     ['43 - Haití 🇭🇹', '44 - Catar 🇶🇦']),
    ('Rafael Ortiz', 'Sevilla', 'rafael.ortiz@email.com',
     ['1 - Brasil 🇧🇷', '2 - España 🇪🇸', '5 - Francia 🇫🇷'],
     ['9 - Noruega 🇳🇴', '11 - Colombia 🇨🇴'],
     ['17 - Senegal 🇸🇳', '20 - EEUU 🇺🇸', '21 - Turquía 🇹🇷', '25 - Marruecos 🇲🇦'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '38 - Nueva Zelanda 🇳🇿'],
     ['39 - Curaçao 🇨🇼', '46 - Panamá 🇵🇦']),
    ('Inés Cabrera', 'Sevilla', 'ines.cabrera@email.com',
     ['3 - Argentina 🇦🇷', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '8 - Portugal 🇵🇹'],
     ['10 - Bélgica 🇧🇪', '13 - Canadá 🇨🇦'],
     ['18 - Japón 🇯🇵', '23 - Croacia 🇭🇷', '24 - Uruguay 🇺🇾', '28 - Australia 🇦🇺'],
     ['34 - Ghana 🇬🇭', '37 - Costa de Marfil 🇨🇮'],
     ['40 - Jordania 🇯🇴', '45 - Irak 🇮🇶']),
    ('Antonio Díaz', 'Málaga', 'antonio.diaz@email.com',
     ['2 - España 🇪🇸', '5 - Francia 🇫🇷', '6 - Países Bajos 🇳🇱'],
     ['9 - Noruega 🇳🇴', '11 - Colombia 🇨🇴'],
     ['17 - Senegal 🇸🇳', '18 - Japón 🇯🇵', '23 - Croacia 🇭🇷', '25 - Marruecos 🇲🇦'],
     ['34 - Ghana 🇬🇭', '38 - Nueva Zelanda 🇳🇿'],
     ['43 - Haití 🇭🇹', '44 - Catar 🇶🇦']),
    ('Rocío Márquez', 'Málaga', 'rocio.marquez@email.com',
     ['1 - Brasil 🇧🇷', '3 - Argentina 🇦🇷', '7 - Alemania 🇩🇪'],
     ['10 - Bélgica 🇧🇪', '13 - Canadá 🇨🇦'],
     ['20 - EEUU 🇺🇸', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾', '28 - Australia 🇦🇺'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '37 - Costa de Marfil 🇨🇮'],
     ['39 - Curaçao 🇨🇼', '46 - Panamá 🇵🇦']),
    ('Jorge Peña', 'Murcia', 'jorge.pena@email.com',
     ['2 - España 🇪🇸', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '8 - Portugal 🇵🇹'],
     ['9 - Noruega 🇳🇴', '14 - Suiza 🇨🇭'],
     ['17 - Senegal 🇸🇳', '20 - EEUU 🇺🇸', '23 - Croacia 🇭🇷', '25 - Marruecos 🇲🇦'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '37 - Costa de Marfil 🇨🇮'],
     ['42 - Uzbekistán 🇺🇿', '44 - Catar 🇶🇦']),
    ('Sofía Vidal', 'Alicante', 'sofia.vidal@email.com',
     ['1 - Brasil 🇧🇷', '5 - Francia 🇫🇷', '7 - Alemania 🇩🇪'],
     ['10 - Bélgica 🇧🇪', '13 - Canadá 🇨🇦'],
     ['18 - Japón 🇯🇵', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾', '28 - Australia 🇦🇺'],
     ['34 - Ghana 🇬🇭', '38 - Nueva Zelanda 🇳🇿'],
     ['39 - Curaçao 🇨🇼', '46 - Panamá 🇵🇦']),
    ('Pablo Lozano', 'Zaragoza', 'pablo.lozano@email.com',
     ['2 - España 🇪🇸', '3 - Argentina 🇦🇷', '6 - Países Bajos 🇳🇱'],
     ['9 - Noruega 🇳🇴', '11 - Colombia 🇨🇴'],
     ['17 - Senegal 🇸🇳', '20 - EEUU 🇺🇸', '23 - Croacia 🇭🇷', '25 - Marruecos 🇲🇦'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '37 - Costa de Marfil 🇨🇮'],
     ['41 - Sudáfrica 🇿🇦', '44 - Catar 🇶🇦']),
    ('Nuria Medina', 'Granada', 'nuria.medina@email.com',
     ['1 - Brasil 🇧🇷', '4 - Inglaterra 🏴󠁧󠁢󠁥󠁮󠁧󠁿', '8 - Portugal 🇵🇹'],
     ['10 - Bélgica 🇧🇪', '13 - Canadá 🇨🇦'],
     ['18 - Japón 🇯🇵', '21 - Turquía 🇹🇷', '24 - Uruguay 🇺🇾', '28 - Australia 🇦🇺'],
     ['34 - Ghana 🇬🇭', '38 - Nueva Zelanda 🇳🇿'],
     ['40 - Jordania 🇯🇴', '46 - Panamá 🇵🇦']),
    ('Ander Goikoetxea', 'Bilbao', 'ander.goikoetxea@email.com',
     ['2 - España 🇪🇸', '5 - Francia 🇫🇷', '7 - Alemania 🇩🇪'],
     ['9 - Noruega 🇳🇴', '11 - Colombia 🇨🇴'],
     ['17 - Senegal 🇸🇳', '20 - EEUU 🇺🇸', '23 - Croacia 🇭🇷', '25 - Marruecos 🇲🇦'],
     ['33 - Escocia 🏴󠁧󠁢󠁳󠁣󠁴󠁿', '37 - Costa de Marfil 🇨🇮'],
     ['39 - Curaçao 🇨🇼', '45 - Irak 🇮🇶']),
]

# ── Knockout results ───────────────────────────────────────
KNOCKOUT_RESULTS = {
    # R32
    'r32-1': ('México', 'Paraguay', 2, 0, '2026-06-28T19:00:00Z'),
    'r32-2': ('Países Bajos', 'Ghana', 3, 0, '2026-06-29T17:00:00Z'),
    'r32-3': ('Brasil', 'Costa de Marfil', 4, 0, '2026-06-29T20:30:00Z'),
    'r32-4': ('Inglaterra', 'Escocia', 2, 0, '2026-06-30T01:00:00Z'),
    'r32-5': ('Alemania', 'República Checa', 3, 0, '2026-06-30T17:00:00Z'),
    'r32-6': ('Argentina', 'Rep. Dem. del Congo', 2, 0, '2026-06-30T21:00:00Z'),
    'r32-7': ('Bélgica', 'Senegal', 2, 1, '2026-07-01T01:00:00Z'),
    'r32-8': ('Canadá', 'Argelia', 1, 0, '2026-07-01T16:00:00Z'),
    'r32-9': ('España', 'Uruguay', 2, 0, '2026-07-01T20:00:00Z'),
    'r32-10': ('Turquía', 'Portugal', 2, 1, '2026-07-02T00:00:00Z'),
    'r32-11': ('Francia', 'Austria', 3, 1, '2026-07-02T19:00:00Z'),
    'r32-12': ('Colombia', 'Croacia', 2, 1, '2026-07-02T23:00:00Z'),
    'r32-13': ('Corea del Sur', 'Suiza', 0, 1, '2026-07-03T03:00:00Z'),
    'r32-14': ('Marruecos', 'EEUU', 1, 2, '2026-07-03T18:00:00Z'),
    'r32-15': ('Ecuador', 'Japón', 0, 1, '2026-07-03T22:00:00Z'),
    'r32-16': ('Egipto', 'Noruega', 0, 2, '2026-07-04T01:30:00Z'),
    # R16
    'r16-1': ('México', 'Países Bajos', 1, 2, '2026-07-04T17:00:00Z'),
    'r16-2': ('Brasil', 'Inglaterra', 3, 1, '2026-07-04T21:00:00Z'),
    'r16-3': ('Alemania', 'Argentina', 2, 1, '2026-07-05T20:00:00Z'),
    'r16-4': ('Bélgica', 'Canadá', 1, 0, '2026-07-06T00:00:00Z'),
    'r16-5': ('España', 'Turquía', 3, 0, '2026-07-06T19:00:00Z'),
    'r16-6': ('Francia', 'Colombia', 2, 0, '2026-07-07T00:00:00Z'),
    'r16-7': ('Suiza', 'EEUU', 1, 1, '2026-07-07T16:00:00Z'),
    'r16-8': ('Japón', 'Noruega', 1, 0, '2026-07-07T20:00:00Z'),
    # QF
    'qf-1': ('Países Bajos', 'Brasil', 1, 2, '2026-07-09T20:00:00Z'),
    'qf-2': ('Alemania', 'Bélgica', 1, 0, '2026-07-10T19:00:00Z'),
    'qf-3': ('España', 'Francia', 2, 1, '2026-07-11T21:00:00Z'),
    'qf-4': ('Suiza', 'Japón', 1, 0, '2026-07-12T01:00:00Z'),
    # SF
    'sf-1': ('Brasil', 'Alemania', 3, 2, '2026-07-14T19:00:00Z'),
    'sf-2': ('España', 'Suiza', 2, 0, '2026-07-15T19:00:00Z'),
    # 3rd & Final
    '3rd-1': ('Alemania', 'Suiza', 3, 1, '2026-07-18T21:00:00Z'),
    'final-1': ('Brasil', 'España', 2, 1, '2026-07-19T19:00:00Z'),
}

R32_EXCEL_IDS = [537417, 537423, 537415, 537418, 537424, 537416, 537425, 537426, 537422, 537421, 537420, 537419, 537429, 537428, 537427, 537430]
R32_ORDER = ['r32-1','r32-2','r32-3','r32-4','r32-5','r32-6','r32-7','r32-8','r32-9','r32-10','r32-11','r32-12','r32-13','r32-14','r32-15','r32-16']
R16_EXCEL_IDS = [537376, 537375, 537377, 537378, 537379, 537380, 537381, 537382]
R16_ORDER = ['r16-1','r16-2','r16-3','r16-4','r16-5','r16-6','r16-7','r16-8']
QF_EXCEL_IDS = [537383, 537384, 537385, 537386]
QF_ORDER = ['qf-1','qf-2','qf-3','qf-4']
SF_EXCEL_IDS = [537387, 537388]
SF_ORDER = ['sf-1','sf-2']
FINAL_EXCEL_IDS = [537389, 537390]
FINAL_ORDER = ['3rd-1','final-1']

# ── Utils ──────────────────────────────────────────────────
def fmt(team_list):
    return ', '.join(team_list) if team_list else None

# ── Standing calculation ───────────────────────────────────
def calc_standings(matches):
    stats = {t: {'pts': 0, 'gf': 0, 'gc': 0, 'w': 0, 'd': 0, 'l': 0}
             for teams in GRUPOS_MUNDIAL.values() for t in teams}
    for m in matches:
        h, a = m['local'], m['visitante']
        hg, ag = int(m['goles_local'] or 0), int(m['goles_visitante'] or 0)
        if h in stats and a in stats:
            stats[h]['gf'] += hg; stats[h]['gc'] += ag
            stats[a]['gf'] += ag; stats[a]['gc'] += hg
            if hg > ag: stats[h]['pts'] += 3; stats[h]['w'] += 1; stats[a]['l'] += 1
            elif hg < ag: stats[a]['pts'] += 3; stats[a]['w'] += 1; stats[h]['l'] += 1
            else: stats[h]['pts'] += 1; stats[a]['pts'] += 1; stats[h]['d'] += 1; stats[a]['d'] += 1
    standings = {}
    for gid, teams in GRUPOS_MUNDIAL.items():
        ranked = sorted(teams, key=lambda t: (-stats[t]['pts'], -(stats[t]['gf'] - stats[t]['gc']), -stats[t]['gf']))
        standings[gid] = [(t, stats[t]) for t in ranked]
    return stats, standings

# ── Main ───────────────────────────────────────────────────
def main():
    print("📊 Generating group matches...")
    # Build group match list from GROUP_MATCHES dict
    group_matches = []
    for gid in sorted(GRUPOS_MUNDIAL.keys()):
        mids = GROUP_MATCH_IDS[gid]
        for i, kid in enumerate(sorted(k for k in GROUP_MATCHES if k[0] == gid)):
            key = f'{gid}{kid[1:]}'
            home, away, hg, ag, hr, ar, fecha = GROUP_MATCHES[kid]
            match = {
                'matchId': mids[i] if i < len(mids) else 0,
                'local': home, 'visitante': away,
                'goles_local': hg, 'goles_visitante': ag,
                'status': 'FT', 'fase': 'group', 'fecha': fecha,
                'rojas_local': hr, 'rojas_visitante': ar,
            }
            group_matches.append(match)

    # Compute standings from group matches
    stats, standings = calc_standings(group_matches)

    # Get qualifiers
    winners, runners_up, third_placed = [], [], []
    for gid, ranking in standings.items():
        winners.append(ranking[0][0])
        runners_up.append(ranking[1][0])
        third_placed.append((gid, ranking[2][0], ranking[2][1]))

    third_placed.sort(key=lambda x: (-x[2]['pts'], -(x[2]['gf'] - x[2]['gc']), -x[2]['gf']))
    best_third = [t[1] for t in third_placed[:8]]

    # Print standings
    print("\n📊 Group Standings:")
    for gid in sorted(GRUPOS_MUNDIAL.keys()):
        ranking = standings[gid]
        print(f"\n  Grupo {gid}:")
        for i, (team, s) in enumerate(ranking):
            gd = s['gf'] - s['gc']
            qual = ''
            if i == 0: qual = ' 🥇 1º'
            elif i == 1: qual = ' 🥈 2º'
            elif i == 2: qual = ' 🥉 3º' if team in best_third else '  3º'
            else: qual = '  4º'
            print(f"    {i+1}. {team:<20} {s['pts']}pts {s['w']}V {s['d']}E {s['l']}D {s['gf']}:{s['gc']} ({gd:+d}){qual}")

    print(f"\n🏆 Qualifiers: {len(winners)} winners, {len(runners_up)} runners-up, {len(best_third)} best 3rds")
    print(f"   Winners: {', '.join(winners)}")
    print(f"   Runners-up: {', '.join(runners_up)}")
    print(f"   Best 3rds: {', '.join(best_third)}")

    # ── Create workbook ────────────────────────────────────
    wb = openpyxl.Workbook()

    # Sheet 1: Respuestas
    ws = wb.active
    ws.title = 'Respuestas'
    headers_resp = [
        'Marca temporal', 'Correo', 'Nombre completo', 'Provincia',
        'Grupo 1 - Tu elección', 'Grupo 2 · Tus 2 elecciones',
        'Grupo 3 · Tus 4 elecciones', 'Grupo 4 · Tus 3 elecciones',
        'Grupo 5 · Tus 3 elecciones'
    ]
    ws.append(headers_resp)

    base_time = datetime(2026, 5, 28, 17, 0, 0)
    all_participants = ORIGINAL_PARTICIPANTS + NEW_PARTICIPANTS
    for i, (name, prov, email, g1, g2, g3, g4, g5) in enumerate(all_participants):
        ts = base_time + timedelta(minutes=10 * i)
        ws.append([ts.strftime('%Y-%m-%d %H:%M:%S'), email, name, prov,
                   fmt(g1), fmt(g2), fmt(g3), fmt(g4), fmt(g5)])

    # Sheet 2: Resultados
    ws2 = wb.create_sheet('Resultados')
    headers_res = ['matchId', 'local', 'visitante', 'goles_local', 'goles_visitante',
                   'status', 'fase', 'fecha', 'rojas_local', 'rojas_visitante']
    ws2.append(headers_res)

    # Write group matches
    for m in group_matches:
        ws2.append([m['matchId'], m['local'], m['visitante'],
                    m['goles_local'], m['goles_visitante'], m['status'],
                    m['fase'], m['fecha'], m['rojas_local'], m['rojas_visitante']])

    # Write knockout matches
    def write_knockout(mids, order, label):
        for mid, key in zip(mids, order):
            home, away, hg, ag, fecha = KNOCKOUT_RESULTS[key]
            ws2.append([mid, home, away, hg, ag, 'FT', label, fecha, 0, 0])

    write_knockout(R32_EXCEL_IDS, R32_ORDER, 'r32')
    write_knockout(R16_EXCEL_IDS, R16_ORDER, 'r16')
    write_knockout(QF_EXCEL_IDS, QF_ORDER, 'qf')
    write_knockout(SF_EXCEL_IDS, SF_ORDER, 'sf')
    write_knockout(FINAL_EXCEL_IDS[:1], FINAL_ORDER[:1], '3rd')
    write_knockout(FINAL_EXCEL_IDS[1:], FINAL_ORDER[1:], 'final')

    # Sheets 3-5
    ws3 = wb.create_sheet('apuestas')
    ws3.append(['token', 'nombre', 'apellido', 'alias', 'paises', 'creado_en', 'pagado'])
    ws4 = wb.create_sheet('partidos')
    ws4.append(['id', 'local_n', 'visitante_n', 'fase', 'goles_local', 'goles_visitante', 'tarjetas_rojas', 'fecha'])
    ws5 = wb.create_sheet('config')
    ws5.append(['clave', 'valor'])
    ws5.append(['fecha_cierre', '2026-06-11T14:00:00-05:00'])

    # Save
    wb.save(EXCEL_PATH)
    print(f"\n✅ Excel saved to {EXCEL_PATH}")

    # Verify
    wb2 = openpyxl.load_workbook(EXCEL_PATH)
    def count_data_rows(sname):
        ws = wb2[sname]
        return sum(1 for i, row in enumerate(ws.iter_rows(values_only=True)) if i > 0 and row[0] is not None)
    n_parts = count_data_rows('Respuestas')
    n_matches = count_data_rows('Resultados')
    ft_count = sum(1 for i, row in enumerate(wb2['Resultados'].iter_rows(values_only=True)) if i > 0 and row[5] == 'FT')
    # Verify R32 uniqueness
    r32_teams = []
    for row in wb2['Resultados'].iter_rows(values_only=True):
        if row[6] == 'r32' and row[1] and row[2]:
            r32_teams.extend([row[1], row[2]])
    from collections import Counter
    dupes = {k: v for k, v in Counter(r32_teams).items() if v > 1}
    r32_unique = f"⚠️ Duplicates: {dupes}" if dupes else f"{len(set(r32_teams))}/32 unique ✓"

    print(f"\n📋 Final summary:")
    print(f"   Participants: {n_parts}")
    print(f"   Total matches: {n_matches}")
    print(f"   FT matches: {ft_count}")
    print(f"   R32 teams: {r32_unique}")

if __name__ == '__main__':
    main()
